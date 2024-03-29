from typing import Any
from django.shortcuts import render
from openpyxl import Workbook, load_workbook
from rest_framework import generics, status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from users.auth_service import user_auth


from .models import MonthlyRoster, ShiftType, WorkDay
from .serializers import (
    BatchCreateSerializer,
    BatchSerializer,
    ClientShiftTypeSerializer,
    RosterCreateSerializer,
    RosterSerializer,
    RosterUpdateSerializer,
    ShiftTypeSerializer,
    WorkdaySerializer,
)

from users.models import Batch, EmployeeBatchUpload
from utils.utils import get_shift_date

# Create your views here.


class BatchListCreateAPIView(generics.ListCreateAPIView):
    queryset = Batch.objects.all()
    serializer_class = BatchSerializer

    def get_serializer_class(self):
        if self.request.method == "POST":
            return BatchCreateSerializer
        return BatchSerializer


class RosterListCreateAPIView(generics.ListCreateAPIView):
    queryset = MonthlyRoster.objects.all().select_related("batch").order_by('-shift_start_date')
    serializer_class = RosterSerializer

    def get_serializer_class(self):
        if self.request.method == "POST":
            return RosterCreateSerializer
        return RosterSerializer

    def create(self, request, *args, **kwargs):
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)

        print("######...: ", request.data)
        modified_request_data = self.modify_request_data(request.data)

        # Pass the modified request data to the serializer
        serializer = self.get_serializer(data=modified_request_data, many=True)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=201, headers=headers)

    def modify_request_data(self, data):
        copy_data = data.copy()
        modified_data = []
        # Iterate over each payload item
        for item in copy_data:
            # Extract common information
            roster_start_date = item.get("start_date")
            group = item.get("group")
            batch = Batch.objects.filter(name__iexact=group).first().id

            # Iterate over shifts
            for w_day, shift_name in item["shifts"].items():
                work_day = WorkDay.objects.filter(day_symbol=w_day).first()
                shift = ShiftType.objects.filter(name__iexact=shift_name).first()
                print("EXTRACTED SHIFT: ", shift)
                shift_start_date, shift_end_date, week_start_date = get_shift_date(
                    work_day.day_symbol, shift, roster_start_date
                )
                new_dict = {
                    "week_start_date": week_start_date,
                    "work_day": work_day.id,
                    "shift": shift.id,
                    "batch": batch,
                    "shift_start_date": shift_start_date,
                    "shift_end_date": shift_end_date,
                }
                modified_data.append(new_dict)
        return modified_data


class RosterRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = MonthlyRoster.objects.all()
    serializer_class = RosterSerializer

    def get_serializer_class(self):
        if self.request.method == "PUT" or self.request.method == "PATCH":
            return RosterUpdateSerializer
        return RosterSerializer

    def patch(self, request, *args, **kwargs):
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)

        print("###...UPDATE REQUEST DATA: ", request.data)
        batch_name = request.data.get("batch")
        shift_name = request.data.get("shift")
        work_day_name = request.data.get("work_day")
        try:
            batch = Batch.objects.get(name__iexact=batch_name)
            shift = ShiftType.objects.get(name__iexact=shift_name)
            work_day = WorkDay.objects.get(day_symbol__iexact=work_day_name)
            # Update request data with primary keys
            request.data["batch"] = batch.pk
            request.data["shift"] = shift.pk
            request.data["work_day"] = work_day.pk
        except (Batch.DoesNotExist, ShiftType.DoesNotExist, WorkDay.DoesNotExist) as e:
            return Response(
                {"error": "Batch, shift, or work day not found"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().patch(request, *args, **kwargs)

@api_view(['GET'])
def get_shift_type_list(request):
    q_set = ShiftType.objects.all()
    serializer = ClientShiftTypeSerializer(q_set, many=True)
    return Response(data=serializer.data, status=status.HTTP_200_OK)

def create_shift_type_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        ws_shift_type = [ws for ws in wb.worksheets if ws.title=="ShiftType"][0]
        imported_shift_type_counter = 0
        skipped_shift_type_counter = 0
        ws_shift_type_columns = ["name", "start_time", "end_time", "duration"]
        shift_type_rows = ws_shift_type.iter_rows(min_row=2)
        for index, shift_type_row in enumerate(shift_type_rows, start=1):
            ws_shift_type_values = [cell.value for cell in shift_type_row[1:]]
            ws_shift_type_dict = dict(zip(ws_shift_type_columns, ws_shift_type_values))
            shift_type_serializer = ShiftTypeSerializer(data=ws_shift_type_dict)
            if shift_type_serializer.is_valid():
                shift_type_serializer.save()
                imported_shift_type_counter += 1
            else:
                skipped_shift_type_counter +=1
    except Exception as ex:
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    
    data = {
            "total_uploaded_shift_type": imported_shift_type_counter,
            "total_skipped_shift_type": skipped_shift_type_counter,
        }
    return Response(data=data, status=status.HTTP_200_OK)

def create_workday_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        ws_workday = [ws for ws in wb.worksheets if ws.title=="WorkDay"][0]
        imported_workday_counter = 0
        skipped_workday_counter = 0
        ws_workday_columns = ["day_symbol", "day_code"]
        workday_rows = ws_workday.iter_rows(min_row=2)
        for index, workday_row in enumerate(workday_rows, start=1):
            workday_values = [cell.value for cell in workday_row[1:]]
            ws_workday_dict = dict(zip(ws_workday_columns, workday_values))
            ws_workday_serializer = WorkdaySerializer(data=ws_workday_dict)
            if ws_workday_serializer.is_valid():
                ws_workday_serializer.save()
                imported_workday_counter += 1
            else:
                skipped_workday_counter +=1
    except Exception as ex:
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    
    data = {
            "total_uploaded_workday": imported_workday_counter,
            "total_skipped_workday": skipped_workday_counter,
        }
    return Response(data=data, status=status.HTTP_200_OK)
