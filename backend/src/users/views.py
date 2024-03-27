from typing import Any
import jwt, datetime

from .field_choices_serializers import (
    BatchField,
    DepartmentField,
    EmployeeCategoryField,
    EmployeeStatusField,
    LocationField,
)
from mealmanager.settings._base import JWT_SALT
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import Workbook, load_workbook
from rest_framework.decorators import api_view, renderer_classes
from rest_framework.views import APIView
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, FileUploadParser
from rest_framework.response import Response
from rest_framework import generics, permissions, status
from django.shortcuts import get_object_or_404
from .serializers import (
    BatchSerializer,
    BatchUploadUserProfileSerializer,
    DepartmentSerializer,
    EmployeeCategorySerializer,
    EmployeeStatusSerializer,
    GetUserFromExcelSerializer,
    GetUserProfileFromExcelSerializer,
    LocationSerializer,
    UserSerializer,
    UserProfileSerializer,
    AvatarSerializer,
)
from .models import (
    Batch,
    Department,
    EmployeeCategory,
    EmployeeStatus,
    Location,
    User,
    UserProfile,
    EmployeeBatchUpload,
)
from .auth_service import user_auth


# Create your views here.
class RegisterView(APIView):
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class LoginView(APIView):
    def post(self, request):
        username = request.data["username"]
        password = request.data["password"]
        print("PASSWORD: ", password)

        user = User.objects.filter(username=username).first()
        if user is None:
            return Response(
                {"invalid_user_error": "User not found!"},
                status=status.HTTP_403_FORBIDDEN,
            )

        if not user.check_password(password):
            # raise AuthenticationFailed("Incorrect password!")
            return Response(
                {"password_error": "Incorrect password!"},
                status=status.HTTP_403_FORBIDDEN,
            )

        if user.is_superuser:
            payload = {
                "id": user.id,
                "exp": datetime.datetime.utcnow() + datetime.timedelta(days=365 * 10),
                "iat": datetime.datetime.utcnow(),
            }
            token = jwt.encode(payload, JWT_SALT, algorithm="HS256")
            print("####...SUPER", token)
            return Response({"jwt": token}, status=status.HTTP_200_OK)
        payload = {
            "id": user.id,
            "exp": datetime.datetime.utcnow() + datetime.timedelta(days=1),
            "iat": datetime.datetime.utcnow(),
        }
        token = jwt.encode(payload, JWT_SALT, algorithm="HS256")
        print("####....OTHERS", token)
        return Response({"jwt": token}, status=status.HTTP_200_OK)


class AuthUserView(APIView):
    def get(self, request, pk=None):
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)
        user = User.objects.filter(id=payload["id"]).first()
        try:
            user_serializer = UserSerializer(user)
            profile = UserProfile.objects.select_related("user").get(user=user.pk)
            profile_serializer = UserProfileSerializer(profile)
            profile_serializer.data["user"] = user_serializer.data
            category_obj = EmployeeCategory.objects.filter(
                id=profile_serializer.data["category"]
            ).first()
            response_data = {
                **profile_serializer.data,
                "category": {
                    "id": category_obj.id,
                    "privilege": category_obj.meal_access,
                },
                "user": {**user_serializer.data},
            }

            response_data["id"] = str(response_data["id"]).replace("-", "")
            return Response(data=response_data, status=status.HTTP_200_OK)
        except UserProfile.DoesNotExist:
            return Response(
                data=user_serializer.data,
                status=status.HTTP_200_OK,
            )


class LogoutView(APIView):
    def post(self, request):
        response = Response()
        response.delete_cookie("jwt")
        response.data = {"message": "success"}
        return response


class UserProfileView(APIView):
    def get(self, request, format=None):
        profiles = UserProfile.objects.all()
        serializer = UserProfileSerializer(profiles, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, format=None):
        print("CREATE PROFILE DATA: ", request.data)
        request_user = request.data.pop("user", None)
        if request_user is None:
            return Response(
                {"error": "User.id field is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)
        auth_user = User.objects.filter(id=payload["id"]).first()
        user_data = {
            "first_name": request_user.pop("first_name", ""),
            "last_name": request_user.pop("last_name", ""),
            "middle_name": request_user.pop("middle_name", ""),
            "email": auth_user.email,
            "username": auth_user.username,
            "password": auth_user.password,
        }
        user_serializer = UserSerializer(instance=auth_user, data=user_data)
        modified_data = {**request.data, "user": request_user["id"]}
        profile_serializer = UserProfileSerializer(data=modified_data)

        if profile_serializer.is_valid() and user_serializer.is_valid():
            user_serializer.save()
            profile_serializer.save()
            category_obj = EmployeeCategory.objects.filter(
                id=profile_serializer.data["category"]
            ).first()
            response_data = {
                **profile_serializer.data,
                "category": {
                    "id": category_obj.id,
                    "privilege": category_obj.meal_access,
                },
                "user": {**user_serializer.data},
            }
            print("###...RESPONSE PROFILE DATA", response_data)
            return Response(response_data, status=status.HTTP_201_CREATED)
        profile_serializer.is_valid()
        user_serializer.is_valid()
        response = Response(
            {
                "errors": {
                    "profile": {**profile_serializer.errors},
                    "user": {**user_serializer.errors},
                }
            },
            status=status.HTTP_400_BAD_REQUEST,
        )
        return response


def get_object_by_lookup_field(model, lookup_field, lookup_value):
    """
    Retrieve an object using the specified lookup field and value.
    """
    lookup_kwargs = {lookup_field: lookup_value}
    return get_object_or_404(model, **lookup_kwargs)


class UserProfileDetailAPIView(APIView):
    lookup_url_kwargs = "id"
    lookup_field = "id"

    def get(self, request, id, format=None):
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)
        user = User.objects.filter(id=payload["id"]).first()

        profile_instance = get_object_by_lookup_field(
            UserProfile, self.lookup_field, id
        )
        user_serializer = UserSerializer(user)
        profile_serializer = UserProfileSerializer(profile_instance)
        category_obj = EmployeeCategory.objects.filter(
            id=profile_serializer.data["category"]
        ).first()
        response_data = {
            **profile_serializer.data,
            "category": {
                "id": category_obj.id,
                "privilege": category_obj.meal_access,
            },
            "user": {**user_serializer.data},
        }
        return Response(data=response_data, status=status.HTTP_200_OK)

    def patch(self, request, id, format=None):
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)
        user = User.objects.filter(id=payload["id"]).first()

        profile_instance = get_object_by_lookup_field(
            UserProfile, self.lookup_field, id
        )
        print("*********")
        print("PROFILE INSTANCE USER: ", profile_instance.user)
        print("USER INSTANCE USER: ", user)
        print("*********")
        if profile_instance.user != user and not user.is_superuser:
            return Response(
                data={"access_error": "You are not allowed to modify profile"},
                status=status.HTTP_403_FORBIDDEN,
            )

        request_user = request.data.pop("user", None)
        if request_user is None:
            return Response(
                {"error": "User.id field is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            auth_user = User.objects.get(pk=request_user["id"])
        except User.DoesNotExist:
            return Response(
                {"errors": f"User with id {request_user['id']} not found!"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user_data = {
            "first_name": request_user.pop("first_name", auth_user.first_name),
            "last_name": request_user.pop("last_name", auth_user.last_name),
            "middle_name": request_user.pop("middle_name", auth_user.middle_name),
            "email": auth_user.email,
            "username": auth_user.username,
            "password": auth_user.password,
        }
        user_serializer = UserSerializer(instance=auth_user, data=user_data)
        if profile_instance is None:
            return Response(
                {"error": f"User profile with id <{id}> not found!"},
                status=status.HTTP_404_NOT_FOUND,
            )
        modified_data = {**request.data, "user": request_user["id"]}
        profile_serializer = UserProfileSerializer(
            instance=profile_instance, data=modified_data
        )
        if profile_serializer.is_valid() and user_serializer.is_valid():
            user_serializer.save()
            profile_serializer.save()
            category_obj = EmployeeCategory.objects.filter(
                id=profile_serializer.data["category"]
            ).first()
            response_data = {
                **profile_serializer.data,
                "category": {
                    "id": category_obj.id,
                    "privilege": category_obj.meal_access,
                },
                "user": {**user_serializer.data},
            }

            return Response(response_data, status=status.HTTP_200_OK)
        profile_serializer.is_valid()
        user_serializer.is_valid()
        return Response(
            {
                "error": {
                    "profile": {**profile_serializer.errors},
                    "user": {**user_serializer.errors},
                }
            },
            status=status.HTTP_400_BAD_REQUEST,
        )


class ProfileFieldChoicesView(APIView):
    def get(self, request, format=None):
        locations = Location.objects.all()  # .values("id", "name")
        print("Location: ", locations)
        departments = Department.objects.all()
        emp_status = EmployeeStatus.objects.all()
        categories = EmployeeCategory.objects.all()
        groups = Batch.objects.all()

        loc_serializer = LocationField(locations, many=True)
        departments_serializer = DepartmentField(departments, many=True)
        emp_status_serializer = EmployeeStatusField(emp_status, many=True)
        categories_serializer = EmployeeCategoryField(categories, many=True)
        group_serializer = BatchField(groups, many=True)

        choice_dict_values = [
            loc_serializer.data,
            departments_serializer.data,
            emp_status_serializer.data,
            categories_serializer.data,
            group_serializer.data,
        ]
        choice_dict_keys = ["location", "department", "emp_status", "category", "group"]
        choice_dict = dict(zip(choice_dict_keys, choice_dict_values))

        return Response(data=choice_dict, status=status.HTTP_200_OK)


class UploadProfileImageView(APIView):
    parser_classes = [FormParser, MultiPartParser, FileUploadParser]

    def patch(self, request, format=None):
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)
        user = User.objects.get(id=payload["id"])
        profile = UserProfile.objects.get(user=user.pk)
        serializer = AvatarSerializer(instance=profile, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)




@api_view(["POST"])
def create_users_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    wb: Workbook = load_workbook(filename=batch_file)
    # ws_user = wb.worksheets[0]
    ws_user = [ws for ws in wb.worksheets if ws.title=="User"][0]
    # ws_profile = wb.worksheets[1]
    ws_profile = [ws for ws in wb.worksheets if ws.title=="Profile"][0]
    imported_user_counter = 0
    skipped_user_counter = 0
    imported_profile_counter = 0
    skipped_profile_counter = 0
    ws_user_columns = [
        "first_name",
        "last_name",
        "middle_name",
        "is_staff",
        "is_active",
        "username",
        "reader_uid",
        "password",
    ]
    ws_profile_columns = [
        "employee_id",
        "reader_uid",
        "gender",
        "category",
        "department",
        "location",
        "employee_status",
        "batch",
        "user",
    ]
    user_rows = ws_user.iter_rows(min_row=2)
    profile_rows = ws_profile.iter_rows(min_row=2)
    for index, user_row in enumerate(user_rows, start=1):
        user_row_values = [cell.value for cell in user_row[1:]]
        user_row_dict = dict(zip(ws_user_columns, user_row_values))
        # print("@@@@ Columns From User: ", user_row_dict.keys())
        # print("@@@@ Values From User: ", user_row_dict.values())
        print("@@@@ Total Columns From User: ", len(user_row_dict.keys()))
        user_serializer = GetUserFromExcelSerializer(data=user_row_dict)
        if user_serializer.is_valid():
            # print("@@@@ VALID USER: ")
            user_serializer.save()
            print("@@@@ VALID USER DATA: ", user_serializer.data)
            imported_user_counter += 1
        else:
            print("@@@@ NOT VALID USER:", user_serializer.errors)
            skipped_user_counter += 1
    
    for index, profile_row in enumerate(profile_rows, start=1):
        profile_row_values = [cell.value for cell in profile_row[1:]]
        # print("#### PROFILE ROLE VALUES: ", profile_row_values)
        reader_uid = profile_row_values[1]
        category_name = profile_row_values[3]
        department_name = profile_row_values[4]
        location_name = profile_row_values[5]
        employee_status_name = profile_row_values[6]
        batch_name = profile_row_values[7]
        print("&&&& READER UID", reader_uid)
        try:
            user = User.objects.get(reader_uid=reader_uid)
            category = EmployeeCategory.objects.get(cat_name=category_name)
            department = Department.objects.get(dept_name=department_name)
            location = Location.objects.get(name=location_name)
            employee_status = EmployeeStatus.objects.get(status=employee_status_name)
            batch = Batch.objects.get(name=batch_name)
            
            profile_row_values[3] = category.pk
            profile_row_values[4] = department.pk
            profile_row_values[5] = location.pk
            profile_row_values[6] = employee_status.pk
            profile_row_values[7] = batch.pk
            profile_row_values.append(user.pk)
            print("#### PROFILE ROLE VALUES_2: ", profile_row_values)
            
            profile_row_dict = dict(zip(ws_profile_columns, profile_row_values))
            print("%%%%%%%%- ROW DICT:", profile_row_dict)
            p_serializer = GetUserProfileFromExcelSerializer(data=profile_row_dict)
            if p_serializer.is_valid():
                p_serializer.save()
                print("@@@@  VALID PROFILE: ", p_serializer.data)
                imported_profile_counter += 1
            else:
                print("@@@@ NOT VALID PROFILE:", p_serializer.errors)
                skipped_profile_counter += 1
        
        except Exception as ex:
            print("Exception occurred: ", ex.args[0])
            return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    data = (
        {
            "total_uploaded_users": imported_user_counter,
            "total_uploaded_profiles": imported_profile_counter,
            "total_skipped_users": skipped_user_counter,
            "total_skipped_profiles": skipped_profile_counter,
        },
    )
    return Response(data=data, status=status.HTTP_200_OK)


def create_location_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        # ws_location = wb.worksheets[2]
        ws_location = [ws for ws in wb.worksheets if ws.title=="Location"][0]
        imported_location_counter = 0
        skipped_location_counter = 0
        ws_location_columns = ["name", "address"]
        location_rows = ws_location.iter_rows(min_row=2)
        for index, location_row in enumerate(location_rows, start=1):
            location_rows_values = [cell.value for cell in location_row[1:]]
            location_row_dict = dict(zip(ws_location_columns, location_rows_values))
            loc_serializer = LocationSerializer(data=location_row_dict)
            if loc_serializer.is_valid():
                loc_serializer.save()
                imported_location_counter += 1
            else:
                skipped_location_counter += 1
        
        data = (
            {
                "total_uploaded_location": imported_location_counter,
                "total_skipped_location": skipped_location_counter,
            },
        )
        return Response(data=data, status=status.HTTP_200_OK)
    except Exception as ex:
        print("Exception occurred: ", ex.args[0])
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)


def create_status_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        # ws_emp_status = wb.worksheets[3]
        ws_emp_status = [ws for ws in wb.worksheets if ws.title=="EmployeeStatus"][0]
        imported_emp_status_counter = 0
        skipped_emp_status_counter = 0
        ws_emp_status_columns = ["status", "description"]
        emp_status_rows = ws_emp_status.iter_rows(min_row=2)
        for index, emp_status_row in enumerate(emp_status_rows, start=1):
            emp_status_values = [cell.value for cell in emp_status_row[1:]]
            emp_status_row_dict = dict(zip(ws_emp_status_columns, emp_status_values))
            emp_status_serializer = EmployeeStatusSerializer(data=emp_status_row_dict)
            if emp_status_serializer.is_valid():
                emp_status_serializer.save()
                imported_emp_status_counter += 1
            else:
                skipped_emp_status_counter += 1
    except Exception as ex:
        print("Exception occurred: ", ex.args[0])
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    
    data = (
        {
            "total_uploaded_emp_status": imported_emp_status_counter,
            "total_skipped_emp_status": skipped_emp_status_counter,
        },
    )
    return Response(data=data, status=status.HTTP_200_OK)

def create_category_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        # ws_category = wb.worksheets[4]
        ws_category = [ws for ws in wb.worksheets if ws.title=="Category"][0]
        imported_category_counter = 0
        skipped_category_counter = 0
        ws_category_columns = ["cat_name", "meal_access", "drink_access", "description"]
        category_rows = ws_category.iter_rows(min_row=2)
        for index, category_row in enumerate(category_rows, start=1):
            category_row_values = [cell.value for cell in category_row[1:]]
            category_row_dict = dict(zip(ws_category_columns, category_row_values))
            category_serializer = EmployeeCategorySerializer(data=category_row_dict)
            if category_serializer.is_valid():
                category_serializer.save()
                imported_category_counter += 1
            else:
                skipped_category_counter += 1
    except Exception as ex:
        print("Exception occurred: ", ex.args[0])
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    
    data = (
        {
            "total_uploaded_category": imported_category_counter,
            "total_skipped_category": skipped_category_counter,
        },
    )
    return Response(data=data, status=status.HTTP_200_OK)

def create_department_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        # ws_department = wb.worksheets[5]
        ws_department = [ws for ws in wb.worksheets if ws.title=="Department"][0]
        imported_department_counter = 0
        skipped_department_counter = 0
        ws_department_columns = ["dept_name"]
        department_rows = ws_department.iter_rows(min_row=2)
        for index, department_row in enumerate(department_rows, start=1):
            department_row_values = [cell.value for cell in department_row[1:]]
            department_row_dict = dict(zip(ws_department_columns, department_row_values))
            dept_serializer = DepartmentSerializer(data=department_row_dict)
            if dept_serializer.is_valid():
                dept_serializer.save()
                imported_department_counter += 1
            else:
                skipped_department_counter += 1
    except Exception as ex:
        print("Exception occurred: ", ex.args[0])
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    
    data = (
        {
            "total_uploaded_department": imported_department_counter,
            "total_skipped_department": skipped_department_counter,
        },
    )
    return Response(data=data, status=status.HTTP_200_OK)

def create_group_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        # ws_department = wb.worksheets[5]
        ws_group = [ws for ws in wb.worksheets if ws.title=="Group"][0]
        imported_group_counter = 0
        skipped_group_counter = 0
        ws_group_columns = ["name"]
        group_rows = ws_group.iter_rows(min_row=2)
        for index, group_row in enumerate(group_rows, start=1):
            group_row_values = [cell.value for cell in group_row[1:]]
            group_row_dict = dict(zip(ws_group_columns, group_row_values))
            group_serializer = BatchSerializer(data=group_row_dict)
            if group_serializer.is_valid():
                group_serializer.save()
                imported_group_counter += 1
            else:
                skipped_group_counter += 1
    except Exception as ex:
        print("Exception occurred: ", ex.args[0])
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    
    data = {
            "total_uploaded_group": imported_group_counter,
            "total_skipped_group": skipped_group_counter,
        }
    
    return Response(data=data, status=status.HTTP_200_OK)
