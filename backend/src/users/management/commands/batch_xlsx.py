from django.core.management.base import BaseCommand


class Command(BaseCommand):
    help = ("Import excel from local .xlsx file.")

    SILENT, NORMAL, VERBOSE, VERY_VERBOSE = 0, 1, 2, 3

    def add_arguments(self, parser):
        # Positional arguments
        parser.add_argument("file_path",
                                 nargs=1,
                                 type=str)

    def handle(self, *args, **options):
        self.verbosity = options.get("verbosity", self.NORMAL) 
        self.file_path = options["file_path"][0] 
        self.prepare()
        self.main()
        self.finalize()

    def prepare(self): 
        self.imported_counter = 0 
        self.skipped_counter = 0


    def main(self):
        import json
        from openpyxl import load_workbook
        from ...serializers import  UserSerializer


        wb = load_workbook(filename=self.file_path)
        ws = wb.worksheets[0]

        if self.verbosity >= self.NORMAL:
            self.stdout.write("=== import users ===")
        columns = ["first_name","last_name", "is_staff", "is_active",
                   "username","email","password"]
        rows = ws.iter_rows(min_row=2)
        for  index, row in enumerate(rows, start=1):
            row_values = [cell.value for cell in row[1:]]
            row_dict = dict(zip(columns, row_values))
            for cell in row[1:]:
                if cell.col_idx == 7:
                    print(f"Column is: {cell.col_idx} and Value is: {cell.value}\n")
            # self.stdout.write(f"@@@@ Row Dict: {row_dict} @@@@@\n")
            serializer = UserSerializer(data=row_dict)
            if serializer.is_valid():
                serializer.save()
                # print("Excel Rows: ",serializer.data)
                if self.verbosity >= self.NORMAL:
                    self.stdout.write(f"Excel Rows: {serializer.data}\n")
                    self.imported_counter += 1
            else:
                if self.verbosity >= self.NORMAL:
                    self.stderr.write(
                                f"Errors importing users"
                        f"{row_dict['first_name']} -"
                                 f"{row_dict['last_name']}:\n"
                    )
                    self.stderr.write(f"{json.dumps(serializer.errors)}\n")
                self.skipped_counter += 1

    def finalize(self):
        if self.verbosity >= self.NORMAL:
            self.stdout.write(f"-------------------------\n")
            self.stdout.write(f"Users imported: {self.imported_counter}\n")
            self.stdout.write(f"Users skipped: {self.skipped_counter}\n\n")

