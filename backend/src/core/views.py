from typing import Any
import jwt, datetime
import json
from django.conf import settings
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q, Sum
from openpyxl import Workbook, load_workbook
from rest_framework.decorators import api_view, renderer_classes
from rest_framework.exceptions import NotFound
from rest_framework.renderers import JSONRenderer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics, permissions, status
from rest_framework.exceptions import AuthenticationFailed

from .paginations import CustomPagination

from .transaction_event_handler import (
    smartcard_handler_for_bar,
    smartcard_handler_for_restaurant,
)
from .serializers import (
    DrinkCategoryListSerializer,
    DrinkCategorySerializer,
    DrinkSerializer,
    ExcelDrinkSerializer,
    TransactionReportSerializer,
    TransactionSerializer,
)
from .models import Drink, DrinkCart, DrinkCategory, Transaction
from users.models import EmployeeBatchUpload, User, UserProfile
from users.auth_service import user_auth
from .reports import export_to_excel, export_to_pdf, report, filter_queryset


# Create your views here.


@api_view(["GET"])
def export_transaction_report_data(request):
    exporter = export_to_excel
    if request.query_params.get("exporter") == "pdf":
        exporter = export_to_pdf
    q_set = filter_queryset(request, Transaction)
    serializer = TransactionReportSerializer(q_set, many=True)
    serializer_data = serializer.data
    response = exporter(serializer_data)
    return response


@api_view(["POST"])
def drink_transaction(request):
    request_data = request.data
    print("####...: ", request_data)
    request_json = json.dumps(request_data)
    owner_profile = UserProfile.objects.get(id=request_data["owner_profile"])
    drink_category = owner_profile.category.drink_access
    reader_uid = owner_profile.reader_uid
    # ...get today's transaction
    today = timezone.now().date()
    drink_taken_queryset = DrinkCart.objects.filter(
        reader_uid=reader_uid, order_date__date=today
    )
    first_drink = drink_taken_queryset.first()
    total_drink_taken = 0
    if first_drink:
        total_drink_taken = drink_taken_queryset.aggregate(total_qty=Sum("qty"))[
            "total_qty"
        ]
    print("TAKEN: ", total_drink_taken)
    balance = drink_category - total_drink_taken
    sum_cart = sum([val for key, val in request_data["cart_item"].items()])
    print("Sum Cart: ", sum_cart)
    if (sum_cart > drink_category) or (sum_cart > balance):
        return Response(
            data={"error": "Excess drinks taken"}, status=status.HTTP_400_BAD_REQUEST
        )
    # ...create Transaction
    INCREMENT = 1
    SWIPE_COUNT = int(request_data["swipe_count"]) + INCREMENT
    transaction = Transaction.objects.create(
        owner=owner_profile,
        authorizer=owner_profile,
        swipe_count=SWIPE_COUNT,
        reader_uid=reader_uid,
        access_point=request_data["access_point"],
        raw_payload=request_json,
        grant_type=request_data["grant_type"],
    )
    print("###Transaction: ", transaction)
    serializer = TransactionSerializer(transaction)
    # ...create DrinkCart
    for key, val in request_data["cart_item"].items():
        try:
            drink = Drink.objects.get(drink=key.capitalize())
            DrinkCart.objects.create(
                drink=drink, qty=val, transaction=transaction, reader_uid=reader_uid
            )
        except Drink.DoesNotExist:
            Response(
                data={"error": f"Drink name [{key}]not found!"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    drink_taken_queryset = DrinkCart.objects.filter(
        reader_uid=reader_uid, order_date__date=today
    )
    total_drink_taken = drink_taken_queryset.aggregate(total_qty=Sum("qty"))[
        "total_qty"
    ]
    balance = drink_category - total_drink_taken
    response_data = {
        "balance": balance,
        "allow_access": drink_category,
        "used_access": total_drink_taken,
        "employee": (owner_profile.user.full_name).capitalize(),
    }
    return Response(data=response_data, status=status.HTTP_200_OK)


@api_view(["GET"])
def get_drink_list(request):
    # payload = user_auth(request)
    # if payload.get("auth_error", None):
    #     return Response(payload, status=status.HTTP_403_FORBIDDEN)
    drink_categories = DrinkCategory.objects.all()
    serializer = DrinkCategoryListSerializer(drink_categories, many=True)
    return Response(data=serializer.data, status=status.HTTP_200_OK)


@api_view(["GET", "POST"])
def get_restaurant_transaction_details(request, pk=None):
    payload = user_auth(request)
    if payload.get("auth_error", None):
        return Response(payload, status=status.HTTP_403_FORBIDDEN)
    _user = User.objects.filter(id=payload["id"]).first()
    if not _user.is_superuser or not _user.is_staff:
        return Response(
            data={
                "error": "User is not an Admin!",
            },
            status=status.HTTP_401_UNAUTHORIZED,
        )
    if not request.data:
        return Response(data={"NO PAYLOAD!"}, status=status.HTTP_400_BAD_REQUEST)
    if request.method == "POST":
        ACCESS_GRANTED = "ACCESS GRANTED"
        request_data = {**request.data}
        # get today's transaction based on grant type and uid
        today = timezone.now().date()
        transactions = Transaction.objects.filter(
            reader_uid=request_data["uid"], date__date=today
        )
        if transactions.count() == 0:
            return Response(data={"NO TRANSACTION MATCH!"}, status=status.HTTP_200_OK)
        approved_transaction_counts = transactions.filter(
            grant_type=ACCESS_GRANTED,
        ).count()
        last_transaction = transactions.last()
        serializer = TransactionSerializer(last_transaction)
        owner_profile = UserProfile.objects.get(reader_uid=request_data["uid"])
        meal_category = owner_profile.category.meal_access
        username = owner_profile.user.username
        balance = meal_category - approved_transaction_counts
        avatar = str(owner_profile.profile_image)
        grant_type = serializer.data["grant_type"]
        response_data = {
            "meal_category": meal_category,
            "used_count": approved_transaction_counts,
            "balance": balance,
            "username": username,
            "avatar": avatar,
            "department": str(owner_profile.department),
            "grant_type": grant_type,
            "message": request_data["message"],
        }
        return Response(data=response_data, status=status.HTTP_200_OK)
    return Response(data={_user.username}, status=status.HTTP_200_OK)

    # {"meal_category": 5, "used": 1, "balance": 4, "username": "", "avatar": ""}
    # {"bulk_create": 1}


class TransactionReportApiView(generics.ListCreateAPIView):
    queryset = Transaction.objects.all().select_related()
    serializer_class = TransactionReportSerializer
    pagination_class = CustomPagination  # Apply custom pagination class

    # def get(self, request, *args, **kwargs):
    #     payload = user_auth(self.request)
    #     if "auth_error" in payload.keys():
    #         return Response(payload, status=status.HTTP_403_FORBIDDEN)
    #     return super().get(request, *args, **kwargs)

    def get_queryset(self):
        print("Self REQUEST: ", self.request.query_params)
        q_set = filter_queryset(self.request, Transaction)
        return q_set


class TransactionAPIView(generics.ListCreateAPIView):
    def post(self, request, *args, **kwargs):
        request_data = request.data
        print("OOOOOO====request Data: ", request_data["usb_input"])
        usb_input = request_data["usb_input"]
        if usb_input.startswith("0"):
            usb_input = usb_input[1:]
        try:
            # check if DEPLOYMENT_LOCATION=Restaurant
            if settings.DEPLOYMENT_LOCATION == settings.ACCESS_POINTS["restaurant"]:
                smartcard_handler_for_restaurant(usb_input)
            elif settings.DEPLOYMENT_LOCATION == settings.ACCESS_POINTS["bar"]:
                smartcard_handler_for_bar(usb_input)
        except Exception as ex:
            return Response(data={"error": ex.args[0]})
        return Response(
            data={"message": "Transaction successful!"}, status=status.HTTP_200_OK
        )


class TransactionView(APIView):
    def get_queryset(self):
        report_type = self.request.query_params.get("report-type", None)
        sort_option = self.request.query_params.get("sort-option", None)
        if report_type == "":
            report_type = None
        if sort_option == "":
            sort_option = None
        print("0: @@@@@@@@@@@@@@", report_type, sort_option)
        if report_type is not None and sort_option is not None:
            return report(
                report=report_type.lower(),
                sort_by=sort_option.lower(),
            )
        if report_type is not None and sort_option is None:
            return report(report=report_type.lower())
        if report_type is None and sort_option is not None:
            print("3: @@@@@@@@@@@@@@", report_type, sort_option.lower())
            return report(sort_by=sort_option.lower())
        if report_type is None and sort_option is None:
            print("4@@@@@@@@@@@@@@", report_type, sort_option)
            return report()
        return Transaction.objects.filter().order_by(sort_option)

    def get(self, request, pk=None):
        NOT_ADMIN = "User is not an Admin!"
        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)
        _user = User.objects.filter(id=payload["id"]).first()
        if not _user.is_superuser:
            return Response(
                data={
                    "error": NOT_ADMIN,
                },
                status=status.HTTP_401_UNAUTHORIZED,
            )
        transactions = self.get_queryset()
        serializer = TransactionSerializer(transactions, many=True)
        return Response(data=serializer.data)

    def post(self, request):
        ACCESS_GRANTED = "ACCESS GRANTED"
        ACCESS_DENIED = "ACCESS DENIED"
        ACCESS_POINT = "REMOTE"

        payload = user_auth(request)
        if payload.get("auth_error", None):
            return Response(payload, status=status.HTTP_403_FORBIDDEN)
        _user = User.objects.filter(id=payload["id"]).first()
        if not _user.is_superuser:
            return Response(
                data={
                    "error": "User is not an Admin!",
                },
                status=status.HTTP_401_UNAUTHORIZED,
            )
        try:
            today = timezone.now().date()
            authorizer = UserProfile.objects.get(user=payload["id"])
            request_data = {**request.data}
            card_owner = UserProfile.objects.get(
                user__username=request_data["username"], reader_uid=request_data["uid"]
            )

            today_transactions = Transaction.objects.filter(
                reader_uid=request_data["uid"],
                grant_type=ACCESS_GRANTED,
                date__date=today,
            )
            if today_transactions.first() is None:
                raise ObjectDoesNotExist("No transaction exists yet for this owner")

            owner_profile_data = {
                "id": card_owner.id,
                "user_id": 0,
                "user": {"username": "...", "email": "...", "password": "..."},
                "meal_category": 1,
                # "profile_image": "...",
                "department": "...",
            }
            request_user_profile_data = {
                "id": authorizer.id,
                "user_id": 0,
                "user": {"username": "...", "email": "...", "password": "..."},
                "meal_category": 1,
                # "profile_image": "...",
                "department": "...",
            }
            transaction_count = today_transactions.count()
            meal_category = today_transactions.first().owner.meal_category
            SWIPE_COUNT = transaction_count
            if SWIPE_COUNT < meal_category:
                SWIPE_COUNT += 1
                transaction_data = {
                    "owner": owner_profile_data,
                    "authorizer": request_user_profile_data,
                    "swipe_count": SWIPE_COUNT,
                    "reader_uid": request_data["uid"],
                    "access_point": ACCESS_POINT,
                    "raw_payload": json.dumps(request_data),
                    "grant_type": "ACCESS GRANTED",
                    "owner_id": card_owner.id,
                    "authorizer_id": authorizer.id,
                }

                transaction_serializer = TransactionSerializer(data=transaction_data)
                if transaction_serializer.is_valid():
                    transaction_serializer.save()
                    return Response(data=transaction_serializer.data)
                else:
                    return Response(
                        data=transaction_serializer.errors,
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            if SWIPE_COUNT >= meal_category:
                today_transactions = Transaction.objects.filter(
                    reader_uid=request_data["uid"],
                    date__date=today,
                )
                transaction_count = today_transactions.count()
                NEW_SWIPE_COUNT = transaction_count + 1
                transaction_data = {
                    "owner": owner_profile_data,
                    "authorizer": request_user_profile_data,
                    "swipe_count": NEW_SWIPE_COUNT,
                    "reader_uid": request_data["uid"],
                    "access_point": ACCESS_POINT,
                    "raw_payload": json.dumps(request_data),
                    "grant_type": ACCESS_DENIED,
                    "owner_id": card_owner.id,
                    "authorizer_id": authorizer.id,
                }
                return Response(
                    data={"error": "ACCESS DENIED. YOU HAD ENOUGH MEAL TODAY!"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except User.DoesNotExist:
            raise NotFound("ACCESS DENIED. User not found!")
        except UserProfile.DoesNotExist:
            raise NotFound("ACCESS DENIED. User must setup a profile!")
        except ObjectDoesNotExist:
            authorizer = UserProfile.objects.get(user=payload["id"])
            request_data = {**request.data}
            card_owner = UserProfile.objects.get(
                user__username=request_data["username"]
            )
            owner_profile_data = {
                "id": card_owner.id,
                "user_id": 0,
                "user": {"username": "...", "email": "...", "password": "..."},
                "meal_category": 1,
                # "profile_image": "...",
                "department": "...",
            }
            request_user_profile_data = {
                "id": authorizer.id,
                "user_id": 0,
                "user": {"username": "...", "email": "...", "password": "..."},
                "meal_category": 1,
                # "profile_image": "...",
                "department": "...",
            }
            SWIPE_COUNT = 1
            transaction_data = {
                "owner": owner_profile_data,
                "authorizer": request_user_profile_data,
                "swipe_count": SWIPE_COUNT,
                "reader_uid": request_data["uid"],
                "access_point": ACCESS_POINT,
                "raw_payload": json.dumps(request_data),
                "grant_type": ACCESS_GRANTED,
                "owner_id": card_owner.id,
                "authorizer_id": authorizer.id,
            }

            transaction_serializer = TransactionSerializer(data=transaction_data)
            if transaction_serializer.is_valid():
                transaction_serializer.save()
                return Response(data=transaction_serializer.data)
            else:
                return Response(
                    data=transaction_serializer.errors,
                    status=status.HTTP_400_BAD_REQUEST,
                )

def create_drink_category_from_excel(request):
    try:
        batch_file: Any = EmployeeBatchUpload.objects.get(
            batch_file__icontains="employee_batch"
        ).batch_file
        print("Batch File Found:", batch_file)
    except EmployeeBatchUpload.DoesNotExist:
        return Response(
            data={"error": "No batch file found!"}, status=status.HTTP_400_BAD_REQUEST
        )
    try:
        wb: Workbook = load_workbook(filename=batch_file)
        ws_drink_categ = [ws for ws in wb.worksheets if ws.title=="DrinkCategory"][0]
        ws_drink = [ws for ws in wb.worksheets if ws.title=="Drink"][0]
        imported_drink_categ_counter = 0
        imported_drink_counter = 0
        skipped_drink_categ_counter = 0
        skipped_drink_counter = 0
        ws_drink_categ_columns = ["name"]
        ws_drink_columns = ["drink", "type"]
        drink_categ_rows = ws_drink_categ.iter_rows(min_row=2)
        drink_rows = ws_drink.iter_rows(min_row=2)
        for index, drink_categ_row in enumerate(drink_categ_rows, start=1):
            drink_categ_row_values = [cell.value for cell in drink_categ_row[1:]]
            drink_categ_row_dict = dict(zip(ws_drink_categ_columns, drink_categ_row_values))
            drink_categ_serializer = DrinkCategorySerializer(data=drink_categ_row_dict)
            if drink_categ_serializer.is_valid():
                drink_categ_serializer.save()
                imported_drink_categ_counter += 1
            else:
                skipped_drink_categ_counter += 1
        
        for index, drink_row in  enumerate(drink_rows, start=1):
            drink_row_values = [cell.value for cell in drink_row[1:]]
            categ_name = drink_row_values[1]
            categ = DrinkCategory.objects.get(name=categ_name)
            drink_row_values[1] = categ.pk
            drink_row_dict = dict(zip(ws_drink_columns, drink_row_values))
            drink_serializer = ExcelDrinkSerializer(data=drink_row_dict)
            if drink_serializer.is_valid():
                drink_serializer.save()
                imported_drink_counter += 1
            else:
                skipped_drink_counter += 1
            
    except Exception as ex:
        print("Exception occurred: ", ex.args[0])
        raise
        return Response(data={"error": ex.args[0]}, status=status.HTTP_400_BAD_REQUEST)
    
    data = (
        {
            "total_uploaded_drink_categ": imported_drink_categ_counter,
            "total_skipped_drink_categ": skipped_drink_categ_counter,
            "total_uploaded_drink": imported_drink_counter,
            "total_skipped_drink": skipped_drink_counter,
        },
    )
    return Response(data=data, status=status.HTTP_200_OK)
